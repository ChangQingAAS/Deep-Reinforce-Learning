import numpy as np
from maddpg_action_taking_test import MADDPG
#from maddpg_made.Agent import Agent
from buffer_made import MultiAgentReplayBuffer
from pettingzoo.mpe import simple_adversary_v2
#from pettingzoo.mpe import simple_v2
import openpyxl

PRINT_INTERVAL = 50
N_GAMES = 10000
MAX_STEPS = 50
total_steps = 0
score_history = []
evaluate = False
best_score = 0
test_number = 2

workbook = openpyxl.Workbook()  # workbook = openpyxl.Workbook(encoding='UTF-8')
# 获取活动工作表， 默认就是第一个工作表
worksheet = workbook.active
worksheet.title = "maddpg"

row = ['epoch_number']
for i in range(test_number):
    row.append('average reward%s,' % i)
for i in range(len(row)):
    worksheet.cell(1, i + 1, row[i])


def obs_list_to_state_vector(observation):
    state = np.array([])
    for agent in env.agents:
        state = np.concatenate([state, observation[agent]])
    return state


# scenario is just name for file for saving
#scenario = 'simple_adversary'
env = simple_adversary_v2.parallel_env()
env.reset()

critic_obs_dims = 0
for i in env.agents:
    critic_obs_dims += env.observation_spaces[i].shape[0]

n_actions = []
for i in env.agents:
    n_actions.append(env.action_spaces[i].n)

actor_obs_dims = []
for i in env.agents:
    actor_obs_dims.append(env.observation_spaces[i].shape[0])

maddpg_agents = MADDPG(env, fc1=64, fc2=64, alpha=0.01, beta=0.01, chkpt_dir='testing_save/')

memory = MultiAgentReplayBuffer(env, max_size=1000000, batch_size=1024)

########################## parallel API ################################

if evaluate:
    maddpg_agents.load_checkpoint()

for test_counter in range(test_number):
    count = 2
    for i in range(N_GAMES):
        obs = env.reset()
        #done = env.dones
        done_list = [False for agent in env.agents]

        score = 0
        episode_step = 0

        while not any(done_list):
            #env.render(mode='human')
            actions = maddpg_agents.choose_action(obs, False)
            #print("the actions:", actions)
            obs_, reward, done, infos = env.step(actions)
            done_list = list(done.values())
            if episode_step > MAX_STEPS:
                done_list = [True]  #this might flag an error, might have to move it to after
                #store transition, but that has to move anyway

            state = obs_list_to_state_vector(obs)
            state_ = obs_list_to_state_vector(obs_)

            memory.store_transition(obs, state, actions, reward, obs_, state_, done)

            obs = obs_

            episode_step += 1
            total_steps += 1
            score += sum(reward.values())  # I have a feeling it collects the rewards

            #got to try and get it to learn now
            if total_steps % 100 == 0:
                maddpg_agents.learn(memory)

        score_history.append(score)
        avg_score = np.mean(score_history[-100:])
        #if not evaluate:
        #   if avg_score > best_score:
        #      maddpg_agents.save_checkpoint()
        #     best_score = avg_score
        if i % PRINT_INTERVAL == 0:
            print('episode', i, 'average score {:.1f}'.format(avg_score))
            if test_counter == 0:
                worksheet.cell(count, 1, i)
            worksheet.cell(count, test_counter + 2, avg_score)
            count += 1

env.close()
# print(score)

workbook.save(filename="./result/maddpg.xlsx")
